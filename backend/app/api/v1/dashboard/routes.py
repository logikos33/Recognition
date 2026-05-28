"""
Recognition — Dashboard Routes.

GET /api/v1/dashboard/stats      — System stats overview
GET /api/v1/dashboard/detections — Detection stats by time period
GET /api/v1/reports/export       — Export report as Excel
"""
import io
import logging
from datetime import datetime, timedelta

from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required

from app.core.auth import get_current_user_id
from app.core.exceptions import EpiMonitorError
from app.core.responses import success, error
from app.infrastructure.database.connection import DatabasePool

logger = logging.getLogger(__name__)

dashboard_bp = Blueprint("dashboard", __name__)


def _get_pool() -> DatabasePool:
    pool = DatabasePool.get_instance()
    if pool is None:
        raise RuntimeError("Database pool not initialized")
    return pool


@dashboard_bp.route("/api/v1/dashboard/stats", methods=["GET"])
@jwt_required()
def get_stats():  # type: ignore[no-untyped-def]
    """Overview stats do sistema."""
    try:
        pool = _get_pool()
        stats = {}

        with pool.get_connection() as conn:
            cur = conn.cursor()

            # Cameras
            cur.execute("SELECT COUNT(*) as total FROM cameras")
            stats["cameras_total"] = cur.fetchone()["total"]

            # Videos
            cur.execute("SELECT COUNT(*) as total FROM training_videos")
            stats["videos_total"] = cur.fetchone()["total"]

            cur.execute(
                "SELECT COUNT(*) as total FROM training_videos WHERE status = 'extracted'"
            )
            stats["videos_extracted"] = cur.fetchone()["total"]

            # Frames
            cur.execute("SELECT COUNT(*) as total FROM training_frames")
            stats["frames_total"] = cur.fetchone()["total"]

            cur.execute(
                "SELECT COUNT(*) as total FROM training_frames WHERE is_annotated = TRUE"
            )
            stats["frames_annotated"] = cur.fetchone()["total"]

            # Training jobs
            cur.execute("SELECT COUNT(*) as total FROM training_jobs")
            stats["jobs_total"] = cur.fetchone()["total"]

            cur.execute(
                "SELECT COUNT(*) as total FROM training_jobs WHERE status = 'running'"
            )
            stats["jobs_running"] = cur.fetchone()["total"]

            # Models
            cur.execute("SELECT COUNT(*) as total FROM trained_models")
            stats["models_total"] = cur.fetchone()["total"]

            cur.execute(
                "SELECT COUNT(*) as total FROM trained_models WHERE is_active = TRUE"
            )
            stats["models_active"] = cur.fetchone()["total"]

            # Alerts (last 24h)
            cur.execute(
                "SELECT COUNT(*) as total FROM alerts "
                "WHERE created_at > NOW() - INTERVAL '24 hours'"
            )
            stats["alerts_24h"] = cur.fetchone()["total"]

            cur.execute(
                "SELECT COUNT(*) as total FROM alerts WHERE acknowledged = FALSE"
            )
            stats["alerts_pending"] = cur.fetchone()["total"]

            # Annotations distribution
            cur.execute(
                "SELECT c.name, COUNT(*) as count "
                "FROM frame_annotations a "
                "JOIN yolo_classes c ON a.class_id = c.id "
                "GROUP BY c.name ORDER BY count DESC LIMIT 10"
            )
            stats["class_distribution"] = [
                {"class": row["name"], "count": row["count"]}
                for row in cur.fetchall()
            ]

        return success(stats)
    except EpiMonitorError:
        raise
    except Exception as exc:
        logger.error("get_stats_error: %s", exc, exc_info=True)
        return error("Erro interno", 500)


@dashboard_bp.route("/api/v1/dashboard/detections", methods=["GET"])
@jwt_required()
def get_detection_stats():  # type: ignore[no-untyped-def]
    """Detection stats by day for the last 30 days."""
    try:
        pool = _get_pool()
        days = request.args.get("days", 30, type=int)

        with pool.get_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT DATE(created_at) as day, COUNT(*) as count "
                "FROM alerts "
                "WHERE created_at > NOW() - INTERVAL '%s days' "
                "GROUP BY DATE(created_at) ORDER BY day",
                (days,),
            )
            daily = [
                {"date": str(row["day"]), "count": row["count"]}
                for row in cur.fetchall()
            ]

        return success({"daily_detections": daily, "period_days": days})
    except EpiMonitorError:
        raise
    except Exception as exc:
        logger.error("get_detection_stats_error: %s", exc, exc_info=True)
        return error("Erro interno", 500)


@dashboard_bp.route("/api/v1/reports/export", methods=["GET"])
@jwt_required()
def export_report():  # type: ignore[no-untyped-def]
    """Export alerts report as Excel."""
    try:
        import openpyxl

        pool = _get_pool()
        days = request.args.get("days", 30, type=int)

        with pool.get_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT a.id, c.name as camera_name, a.timestamp, "
                "a.confidence, a.violations, a.acknowledged, a.created_at "
                "FROM alerts a "
                "LEFT JOIN cameras c ON a.camera_id = c.id "
                "WHERE a.created_at > NOW() - INTERVAL '%s days' "
                "ORDER BY a.created_at DESC",
                (days,),
            )
            rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas EPI"

        # Header
        headers = ["ID", "Camera", "Timestamp", "Confianca", "Violacoes", "Reconhecido", "Criado em"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # Data
        for i, row in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=str(row["id"]))
            ws.cell(row=i, column=2, value=row.get("camera_name", "N/A"))
            ws.cell(row=i, column=3, value=str(row["timestamp"]))
            ws.cell(row=i, column=4, value=row["confidence"])
            ws.cell(row=i, column=5, value=str(row.get("violations", [])))
            ws.cell(row=i, column=6, value="Sim" if row["acknowledged"] else "Nao")
            ws.cell(row=i, column=7, value=str(row["created_at"]))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"epi-alertas-{days}d.xlsx",
        )
    except EpiMonitorError:
        raise
    except Exception as exc:
        logger.error("export_report_error: %s", exc, exc_info=True)
        return error("Erro interno", 500)
